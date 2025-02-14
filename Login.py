import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook, Workbook
import os
import sys

# Ensure the Prediction module is in the Python path
sys.path.append(r'C:\Users\kayod\OneDrive\For Uni\Year 4 (Masters)\Final Year Project')

class LoginWindow:
    def __init__(self, master):
        self.master = master
        self.master.title("Login")
        
        self.username_label = ttk.Label(master, text="Username")
        self.username_label.grid(row=0, column=0)
        self.username_entry = ttk.Entry(master)
        self.username_entry.grid(row=0, column=1)
        
        self.password_label = ttk.Label(master, text="Password")
        self.password_label.grid(row=1, column=0)
        self.password_entry = ttk.Entry(master, show="*")
        self.password_entry.grid(row=1, column=1)
        
        self.login_button = ttk.Button(master, text="Login", command=self.check_login)
        self.login_button.grid(row=2, column=0, columnspan=2)
        
        self.create_button = ttk.Button(master, text="Create Account", command=self.create_account)
        self.create_button.grid(row=3, column=0, columnspan=2)

        self.delete_button = ttk.Button(master, text="Delete Account", command=self.delete_account)
        self.delete_button.grid(row=4, column=0, columnspan=2)

        self.message = tk.StringVar()
        self.message_label = ttk.Label(master, textvariable=self.message)
        self.message_label.grid(row=5, column=0, columnspan=2)

    def load_users(self):
        file_path = r'C:\Users\kayod\OneDrive\For Uni\Year 4 (Masters)\Final Year Project\Account Database.xlsx'
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
            users = {sheet.cell(row=i, column=1).value: sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1)}
            return users
        else:
            return {}

    def save_users(self, users):
        file_path = r'C:\Users\kayod\OneDrive\For Uni\Year 4 (Masters)\Final Year Project\Account Database.xlsx'
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Username", "Password"])
        for username, password in users.items():
            sheet.append([username, password])
        workbook.save(file_path)

    def check_login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        
        users = self.load_users()
        if username in users and users[username] == password:
            self.open_prediction_script()
        else:
            self.message.set("Invalid credentials. Please try again.")
    
    def create_account(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        
        users = self.load_users()
        if username in users:
            self.message.set("Username already exists. Please choose another.")
        else:
            users[username] = password
            self.save_users(users)
            self.message.set("Account created successfully.")

    def delete_account(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        
        users = self.load_users()
        if username in users and users[username] == password:
            del users[username]
            self.save_users(users)
            self.message.set("Account deleted successfully.")
        else:
            self.message.set("Invalid credentials. Please try again.")
    
    def open_prediction_script(self):
        self.master.destroy()
        import Prediction  # Import here to ensure it doesn't run on import
        Prediction.run_prediction_app()

if __name__ == "__main__":
    root = tk.Tk()
    login_window = LoginWindow(root)
    root.mainloop()
